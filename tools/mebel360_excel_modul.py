# -*- coding: utf-8 -*-
"""Mebel360° uchun moliyaviy Excel hisobot moduli.

Bu modul app.py ichidan register_excel_module(app, get_db) orqali ulanadi.
"""
from __future__ import annotations

import io
import sqlite3
from datetime import date, datetime, timedelta
from functools import wraps
from typing import Any, Callable

from flask import Blueprint, Response, redirect, render_template_string, request, send_file, session, url_for

try:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, Reference
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError as exc:  # installer openpyxl-ni o‘rnatadi
    raise RuntimeError(
        "Excel moduli uchun openpyxl kerak. CMDda: pip install openpyxl"
    ) from exc


MODULE_NAME = "moliya_excel_v1"
MONEY_FORMAT = '#,##0" so‘m"'
DATE_FORMAT = "yyyy-mm-dd"
THIN = Side(style="thin", color="D9E2F3")
HEADER_FILL = PatternFill("solid", fgColor="166534")
SUBHEADER_FILL = PatternFill("solid", fgColor="DCFCE7")
TITLE_FILL = PatternFill("solid", fgColor="0F172A")
WHITE_FONT = Font(color="FFFFFF", bold=True)
BOLD_FONT = Font(bold=True)


def _table_exists(conn: sqlite3.Connection, table: str) -> bool:
    row = conn.execute(
        "SELECT 1 FROM sqlite_master WHERE type='table' AND name=?",
        (table,),
    ).fetchone()
    return bool(row)


def _columns(conn: sqlite3.Connection, table: str) -> set[str]:
    if not _table_exists(conn, table):
        return set()
    return {str(row[1]) for row in conn.execute(f'PRAGMA table_info("{table}")').fetchall()}


def _safe_iso(value: str | None, fallback: str) -> str:
    text = (value or "").strip()
    try:
        return date.fromisoformat(text).isoformat()
    except ValueError:
        return fallback


def _month_bounds(today: date) -> tuple[str, str]:
    start = today.replace(day=1)
    if today.month == 12:
        next_month = date(today.year + 1, 1, 1)
    else:
        next_month = date(today.year, today.month + 1, 1)
    return start.isoformat(), (next_month - timedelta(days=1)).isoformat()


def _authorized() -> bool:
    # Yangi Mebel360 tizimi.
    if session.get("admin_user_id"):
        role = str(session.get("user_role") or "admin").lower()
        return role in {"admin", "accountant", "hisobchi", "buxgalter"}
    # Eski Mebel360 versiyalari bilan moslik.
    if session.get("logged_in"):
        return True
    if str(session.get("staff_role") or "").lower() in {"hisobchi", "buxgalter", "accountant"}:
        return True
    return False


def _login_redirect() -> Response:
    try:
        return redirect(url_for("login"))
    except Exception:
        return redirect("/login")


def _require_finance_access(view: Callable[..., Any]) -> Callable[..., Any]:
    @wraps(view)
    def wrapped(*args: Any, **kwargs: Any) -> Any:
        if not _authorized():
            return _login_redirect()
        return view(*args, **kwargs)
    return wrapped


def _money(value: Any) -> float:
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def _rows_to_dicts(rows: list[Any]) -> list[dict[str, Any]]:
    return [dict(row) for row in rows]


def _collect_report(
    get_db: Callable[[], sqlite3.Connection],
    start: str,
    end: str,
    order_code: str = "",
    customer: str = "",
    category: str = "",
) -> dict[str, Any]:
    conn = get_db()
    try:
        order_cols = _columns(conn, "buyurtmalar")
        payment_cols = _columns(conn, "buyurtma_tolovlari")
        expense_cols = _columns(conn, "xarajatlar")
        worker_payment_cols = _columns(conn, "tolovlar")
        bonus_cols = _columns(conn, "bonuslar")

        order_where = ["1=1"]
        order_params: list[Any] = []
        if order_code:
            order_where.append("LOWER(b.kod) LIKE LOWER(?)")
            order_params.append(f"%{order_code}%")
        if customer:
            order_where.append("LOWER(b.mijoz) LIKE LOWER(?)")
            order_params.append(f"%{customer}%")
        order_filter = " AND ".join(order_where)

        orders: list[dict[str, Any]] = []
        if order_cols:
            order_date_expr = "date(b.created_at)" if "created_at" in order_cols else "date('now')"
            rows = conn.execute(
                f"""
                SELECT b.*,
                       COALESCE((SELECT SUM(p.miqdor)
                                 FROM buyurtma_tolovlari p
                                 WHERE p.buyurtma_id=b.id),0) AS keyingi_tolovlar
                FROM buyurtmalar b
                WHERE {order_date_expr} BETWEEN ? AND ?
                  AND {order_filter}
                ORDER BY {order_date_expr}, b.id
                """,
                [start, end, *order_params],
            ).fetchall()
            orders = _rows_to_dicts(rows)

        incomes: list[dict[str, Any]] = []

        # Dastlabki avans: buyurtmalar.oldindan_tolov ichidan keyingi to‘lovlarni ayiramiz.
        # Shu tariqa bir pul Excelda ikki marta hisoblanmaydi.
        for row in orders:
            paid_total = _money(row.get("oldindan_tolov"))
            later_total = _money(row.get("keyingi_tolovlar"))
            initial_advance = max(0.0, paid_total - later_total)
            if initial_advance <= 0:
                continue
            created = str(row.get("created_at") or start)[:10]
            incomes.append(
                {
                    "sana": created,
                    "buyurtma_kodi": row.get("kod", ""),
                    "mijoz": row.get("mijoz", ""),
                    "mahsulot": row.get("mahsulot", ""),
                    "turi": "Dastlabki avans",
                    "tolov_usuli": row.get("tolov_usuli", "") if "tolov_usuli" in order_cols else "",
                    "miqdor": initial_advance,
                    "izoh": "Buyurtma ochilganda olingan avans",
                }
            )

        # Keyingi/oraliq/yakuniy to‘lovlar sanasi bo‘yicha olinadi.
        if payment_cols and order_cols:
            pay_method = "p.tolov_usuli" if "tolov_usuli" in payment_cols else "''"
            rows = conn.execute(
                f"""
                SELECT p.sana, b.kod AS buyurtma_kodi, b.mijoz, b.mahsulot,
                       p.turi, {pay_method} AS tolov_usuli,
                       p.miqdor, p.izoh
                FROM buyurtma_tolovlari p
                JOIN buyurtmalar b ON b.id=p.buyurtma_id
                WHERE p.sana BETWEEN ? AND ?
                  AND {order_filter}
                ORDER BY p.sana, p.id
                """,
                [start, end, *order_params],
            ).fetchall()
            incomes.extend(_rows_to_dicts(rows))

        expenses: list[dict[str, Any]] = []
        if expense_cols:
            conditions = ["x.sana BETWEEN ? AND ?"]
            params: list[Any] = [start, end]
            if order_code and "buyurtma_kodi" in expense_cols:
                conditions.append("LOWER(x.buyurtma_kodi) LIKE LOWER(?)")
                params.append(f"%{order_code}%")
            if category and "kategoriya" in expense_cols:
                conditions.append("LOWER(x.kategoriya) LIKE LOWER(?)")
                params.append(f"%{category}%")

            def exp_col(name: str, alias: str | None = None) -> str:
                target = alias or name
                return f"x.{name} AS {target}" if name in expense_cols else f"'' AS {target}"

            rows = conn.execute(
                f"""
                SELECT x.sana,
                       {exp_col("kategoriya")},
                       {exp_col("xarajat_nomi")},
                       {exp_col("buyurtma_kodi")},
                       {exp_col("kimga_berildi")},
                       {exp_col("tolov_usuli")},
                       x.miqdor,
                       {exp_col("tavsifi")},
                       {exp_col("chek_havola")}
                FROM xarajatlar x
                WHERE {' AND '.join(conditions)}
                ORDER BY x.sana, x.id
                """,
                params,
            ).fetchall()
            expenses = _rows_to_dicts(rows)

        worker_payments: list[dict[str, Any]] = []
        if worker_payment_cols:
            worker_join = ""
            worker_select = "'' AS ishchi"
            if _table_exists(conn, "ishchilar") and "ishchi_id" in worker_payment_cols:
                worker_join = "LEFT JOIN ishchilar i ON i.id=t.ishchi_id"
                worker_select = "TRIM(COALESCE(i.ism,'') || ' ' || COALESCE(i.familiya,'')) AS ishchi"
            type_expr = "t.turi" if "turi" in worker_payment_cols else "'To‘lov'"
            desc_expr = "t.tavsifi" if "tavsifi" in worker_payment_cols else "''"
            rows = conn.execute(
                f"""
                SELECT t.sana, {worker_select}, {type_expr} AS turi,
                       t.miqdor, {desc_expr} AS tavsifi
                FROM tolovlar t
                {worker_join}
                WHERE t.sana BETWEEN ? AND ?
                ORDER BY t.sana, t.id
                """,
                (start, end),
            ).fetchall()
            worker_payments = _rows_to_dicts(rows)

        bonuses: list[dict[str, Any]] = []
        if bonus_cols:
            worker_join = ""
            worker_select = "'' AS ishchi"
            if _table_exists(conn, "ishchilar") and "ishchi_id" in bonus_cols:
                worker_join = "LEFT JOIN ishchilar i ON i.id=b.ishchi_id"
                worker_select = "TRIM(COALESCE(i.ism,'') || ' ' || COALESCE(i.familiya,'')) AS ishchi"
            reason_expr = "b.sababi" if "sababi" in bonus_cols else "''"
            rows = conn.execute(
                f"""
                SELECT b.sana, {worker_select}, b.miqdor,
                       {reason_expr} AS sababi
                FROM bonuslar b
                {worker_join}
                WHERE b.sana BETWEEN ? AND ?
                ORDER BY b.sana, b.id
                """,
                (start, end),
            ).fetchall()
            bonuses = _rows_to_dicts(rows)

        total_income = sum(_money(r.get("miqdor")) for r in incomes)
        total_expenses = sum(_money(r.get("miqdor")) for r in expenses)
        total_worker_payments = sum(_money(r.get("miqdor")) for r in worker_payments)
        total_bonuses = sum(_money(r.get("miqdor")) for r in bonuses)
        total_out = total_expenses + total_worker_payments + total_bonuses
        net_profit = total_income - total_out

        initial_advances = sum(
            _money(r.get("miqdor")) for r in incomes if r.get("turi") == "Dastlabki avans"
        )
        later_payments = total_income - initial_advances
        contract_total = sum(_money(r.get("umumiy_narx")) for r in orders)
        paid_total = sum(_money(r.get("oldindan_tolov")) for r in orders)
        debt_total = sum(
            max(0.0, _money(r.get("umumiy_narx")) - _money(r.get("oldindan_tolov")))
            for r in orders
        )

        return {
            "start": start,
            "end": end,
            "orders": orders,
            "incomes": sorted(incomes, key=lambda r: (str(r.get("sana", "")), str(r.get("buyurtma_kodi", "")))),
            "expenses": expenses,
            "worker_payments": worker_payments,
            "bonuses": bonuses,
            "summary": {
                "total_income": total_income,
                "initial_advances": initial_advances,
                "later_payments": later_payments,
                "total_expenses": total_expenses,
                "total_worker_payments": total_worker_payments,
                "total_bonuses": total_bonuses,
                "total_out": total_out,
                "net_profit": net_profit,
                "contract_total": contract_total,
                "paid_total": paid_total,
                "debt_total": debt_total,
                "order_count": len(orders),
            },
        }
    finally:
        conn.close()


def _style_header(ws: Any, row: int, start_col: int, end_col: int) -> None:
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)


def _style_table(ws: Any, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    if max_row < min_row:
        return
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _set_widths(ws: Any, widths: dict[int, float]) -> None:
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _write_empty(ws: Any, row: int, text: str = "Ma’lumot topilmadi") -> int:
    ws.cell(row=row, column=1, value=text)
    ws.cell(row=row, column=1).font = Font(italic=True, color="64748B")
    return row


def _build_workbook(report: dict[str, Any]) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Umumiy hisobot"

    ws.merge_cells("A1:E1")
    ws["A1"] = "MEBEL360° — MOLIYAVIY HISOBOT"
    ws["A1"].fill = TITLE_FILL
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=16)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws["A3"] = "Hisobot davri"
    ws["B3"] = f"{report['start']} — {report['end']}"
    ws["A4"] = "Yaratilgan vaqt"
    ws["B4"] = datetime.now().strftime("%Y-%m-%d %H:%M")
    ws["A5"] = "Izoh"
    ws["B5"] = "Bu boshqaruv hisoboti. Rasmiy soliq/buxgalteriya hisobotini almashtirmaydi."
    ws.merge_cells("B5:E5")
    for cell in ws["A3:A5"]:
        cell[0].font = BOLD_FONT

    summary = report["summary"]
    metrics = [
        ("Jami kirim", summary["total_income"]),
        ("Dastlabki avanslar", summary["initial_advances"]),
        ("Keyingi to‘lovlar", summary["later_payments"]),
        ("Jami xarajat", summary["total_expenses"]),
        ("Ishchi to‘lovlari", summary["total_worker_payments"]),
        ("Bonuslar", summary["total_bonuses"]),
        ("Jami chiqim", summary["total_out"]),
        ("Sof foyda", summary["net_profit"]),
        ("Shartnomalar jami", summary["contract_total"]),
        ("Mijozlardan qolgan qarz", summary["debt_total"]),
        ("Buyurtmalar soni", summary["order_count"]),
    ]
    start_row = 8
    ws.cell(start_row, 1, "Ko‘rsatkich")
    ws.cell(start_row, 2, "Qiymat")
    _style_header(ws, start_row, 1, 2)
    for idx, (label, value) in enumerate(metrics, start=start_row + 1):
        ws.cell(idx, 1, label)
        ws.cell(idx, 2, value)
        ws.cell(idx, 1).font = BOLD_FONT if label in {"Jami kirim", "Jami chiqim", "Sof foyda"} else Font()
        if label != "Buyurtmalar soni":
            ws.cell(idx, 2).number_format = MONEY_FORMAT
        else:
            ws.cell(idx, 2).number_format = "0"
        if label == "Sof foyda":
            ws.cell(idx, 1).fill = SUBHEADER_FILL
            ws.cell(idx, 2).fill = SUBHEADER_FILL
            ws.cell(idx, 2).font = Font(bold=True, color="166534" if value >= 0 else "B91C1C")
    _style_table(ws, start_row + 1, start_row + len(metrics), 1, 2)

    # Grafik uchun ixcham blok
    chart_data_row = start_row + len(metrics) + 3
    chart_rows = [
        ("Kirim", summary["total_income"]),
        ("Xarajat", summary["total_expenses"]),
        ("Ishchi to‘lovi", summary["total_worker_payments"]),
        ("Bonus", summary["total_bonuses"]),
    ]
    for i, (name, value) in enumerate(chart_rows, start=chart_data_row):
        ws.cell(i, 4, name)
        ws.cell(i, 5, value)
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Kirim va chiqimlar"
    chart.y_axis.title = "So‘m"
    chart.height = 7
    chart.width = 12
    data = Reference(ws, min_col=5, min_row=chart_data_row, max_row=chart_data_row + len(chart_rows) - 1)
    cats = Reference(ws, min_col=4, min_row=chart_data_row, max_row=chart_data_row + len(chart_rows) - 1)
    chart.add_data(data, titles_from_data=False)
    chart.set_categories(cats)
    ws.add_chart(chart, "D8")

    _set_widths(ws, {1: 28, 2: 22, 3: 3, 4: 18, 5: 18})
    ws.freeze_panes = "A8"

    # Kirim va to‘lovlar
    ws_income = wb.create_sheet("Kirim va tolovlar")
    income_headers = ["Sana", "Buyurtma kodi", "Mijoz", "Mahsulot", "Turi", "To‘lov usuli", "Miqdor", "Izoh"]
    ws_income.append(income_headers)
    _style_header(ws_income, 1, 1, len(income_headers))
    for row in report["incomes"]:
        ws_income.append([
            row.get("sana", ""),
            row.get("buyurtma_kodi", ""),
            row.get("mijoz", ""),
            row.get("mahsulot", ""),
            row.get("turi", ""),
            row.get("tolov_usuli", ""),
            _money(row.get("miqdor")),
            row.get("izoh", ""),
        ])
    if not report["incomes"]:
        _write_empty(ws_income, 2)
    else:
        for cell in ws_income["G"][1:]:
            cell.number_format = MONEY_FORMAT
    ws_income.auto_filter.ref = f"A1:H{max(2, ws_income.max_row)}"
    ws_income.freeze_panes = "A2"
    _style_table(ws_income, 2, ws_income.max_row, 1, 8)
    _set_widths(ws_income, {1: 13, 2: 16, 3: 24, 4: 24, 5: 20, 6: 18, 7: 18, 8: 36})

    # Xarajatlar
    ws_exp = wb.create_sheet("Xarajatlar")
    exp_headers = ["Sana", "Kategoriya", "Xarajat nomi", "Buyurtma kodi", "Kimga berildi", "To‘lov usuli", "Miqdor", "Tavsifi", "Chek/havola"]
    ws_exp.append(exp_headers)
    _style_header(ws_exp, 1, 1, len(exp_headers))
    for row in report["expenses"]:
        ws_exp.append([
            row.get("sana", ""),
            row.get("kategoriya", ""),
            row.get("xarajat_nomi", ""),
            row.get("buyurtma_kodi", ""),
            row.get("kimga_berildi", ""),
            row.get("tolov_usuli", ""),
            _money(row.get("miqdor")),
            row.get("tavsifi", ""),
            row.get("chek_havola", ""),
        ])
    if not report["expenses"]:
        _write_empty(ws_exp, 2)
    else:
        for cell in ws_exp["G"][1:]:
            cell.number_format = MONEY_FORMAT
    ws_exp.auto_filter.ref = f"A1:I{max(2, ws_exp.max_row)}"
    ws_exp.freeze_panes = "A2"
    _style_table(ws_exp, 2, ws_exp.max_row, 1, 9)
    _set_widths(ws_exp, {1: 13, 2: 20, 3: 24, 4: 16, 5: 22, 6: 18, 7: 18, 8: 36, 9: 32})

    # Buyurtmalar
    ws_orders = wb.create_sheet("Buyurtmalar")
    order_headers = ["Sana", "Kod", "Mijoz", "Mahsulot", "Umumiy narx", "Jami to‘langan", "Qoldiq", "Holat", "Tugash sana"]
    ws_orders.append(order_headers)
    _style_header(ws_orders, 1, 1, len(order_headers))
    for row in report["orders"]:
        total = _money(row.get("umumiy_narx"))
        paid = _money(row.get("oldindan_tolov"))
        ws_orders.append([
            str(row.get("created_at", ""))[:10],
            row.get("kod", ""),
            row.get("mijoz", ""),
            row.get("mahsulot", ""),
            total,
            paid,
            max(0.0, total - paid),
            row.get("holat", ""),
            row.get("tugash_sana", ""),
        ])
    if not report["orders"]:
        _write_empty(ws_orders, 2)
    else:
        for col in ("E", "F", "G"):
            for cell in ws_orders[col][1:]:
                cell.number_format = MONEY_FORMAT
    ws_orders.auto_filter.ref = f"A1:I{max(2, ws_orders.max_row)}"
    ws_orders.freeze_panes = "A2"
    _style_table(ws_orders, 2, ws_orders.max_row, 1, 9)
    _set_widths(ws_orders, {1: 13, 2: 16, 3: 24, 4: 24, 5: 18, 6: 18, 7: 18, 8: 20, 9: 14})

    # Ishchi to‘lovlari
    ws_worker = wb.create_sheet("Ishchi tolovlari")
    worker_headers = ["Sana", "Ishchi", "Turi", "Miqdor", "Tavsifi"]
    ws_worker.append(worker_headers)
    _style_header(ws_worker, 1, 1, len(worker_headers))
    for row in report["worker_payments"]:
        ws_worker.append([
            row.get("sana", ""),
            row.get("ishchi", ""),
            row.get("turi", ""),
            _money(row.get("miqdor")),
            row.get("tavsifi", ""),
        ])
    if not report["worker_payments"]:
        _write_empty(ws_worker, 2)
    else:
        for cell in ws_worker["D"][1:]:
            cell.number_format = MONEY_FORMAT
    ws_worker.auto_filter.ref = f"A1:E{max(2, ws_worker.max_row)}"
    ws_worker.freeze_panes = "A2"
    _style_table(ws_worker, 2, ws_worker.max_row, 1, 5)
    _set_widths(ws_worker, {1: 13, 2: 25, 3: 18, 4: 18, 5: 38})

    # Bonuslar
    ws_bonus = wb.create_sheet("Bonuslar")
    bonus_headers = ["Sana", "Ishchi", "Miqdor", "Sababi"]
    ws_bonus.append(bonus_headers)
    _style_header(ws_bonus, 1, 1, len(bonus_headers))
    for row in report["bonuses"]:
        ws_bonus.append([
            row.get("sana", ""),
            row.get("ishchi", ""),
            _money(row.get("miqdor")),
            row.get("sababi", ""),
        ])
    if not report["bonuses"]:
        _write_empty(ws_bonus, 2)
    else:
        for cell in ws_bonus["C"][1:]:
            cell.number_format = MONEY_FORMAT
    ws_bonus.auto_filter.ref = f"A1:D{max(2, ws_bonus.max_row)}"
    ws_bonus.freeze_panes = "A2"
    _style_table(ws_bonus, 2, ws_bonus.max_row, 1, 4)
    _set_widths(ws_bonus, {1: 13, 2: 25, 3: 18, 4: 42})

    for sheet in wb.worksheets:
        sheet.sheet_view.showGridLines = False

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


PAGE_HTML = r"""
<!doctype html>
<html lang="uz">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Mebel360° — Excel hisobot</title>
<style>
:root{--green:#166534;--light:#f0fdf4;--ink:#0f172a;--muted:#64748b}
*{box-sizing:border-box}body{margin:0;background:#f8fafc;color:var(--ink);font-family:Arial,sans-serif}
.wrap{max-width:1050px;margin:30px auto;padding:0 16px}.card{background:#fff;border:1px solid #e2e8f0;border-radius:18px;padding:22px;box-shadow:0 12px 35px rgba(15,23,42,.08)}
h1{margin:0 0 8px;font-size:28px}.sub{color:var(--muted);margin-bottom:20px}
.grid{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:14px}.full{grid-column:1/-1}
label{display:block;font-weight:700;font-size:13px}input{width:100%;margin-top:6px;padding:12px;border:1px solid #cbd5e1;border-radius:10px;font-size:15px}
.actions{display:flex;gap:10px;flex-wrap:wrap;margin-top:18px}.btn{display:inline-block;border:0;border-radius:11px;padding:12px 16px;font-weight:800;text-decoration:none;cursor:pointer}
.primary{background:var(--green);color:#fff}.secondary{background:#e2e8f0;color:#0f172a}.quick{background:var(--light);color:var(--green);border:1px solid #bbf7d0}
.note{margin-top:18px;padding:13px;border-radius:12px;background:#fff7ed;color:#9a3412;font-size:13px}
@media(max-width:700px){.grid{grid-template-columns:1fr}.full{grid-column:auto}h1{font-size:23px}}
</style>
</head>
<body>
<div class="wrap"><div class="card">
<h1>📊 Moliyaviy hisobotni Excelga chiqarish</h1>
<div class="sub">Kirim, avans, keyingi to‘lovlar, xarajat, ishchi to‘lovi, bonus, qarz va sof foyda.</div>

<div class="actions">
<button class="btn quick" type="button" onclick="preset('today')">Bugun</button>
<button class="btn quick" type="button" onclick="preset('week')">Shu hafta</button>
<button class="btn quick" type="button" onclick="preset('month')">Shu oy</button>
<button class="btn quick" type="button" onclick="preset('last_month')">O‘tgan oy</button>
<button class="btn quick" type="button" onclick="preset('year')">Shu yil</button>
</div>

<form action="{{ export_url }}" method="get">
<div class="grid" style="margin-top:18px">
<label>Boshlanish sana<input id="start" type="date" name="start" value="{{ start }}" required></label>
<label>Tugash sana<input id="end" type="date" name="end" value="{{ end }}" required></label>
<label>Buyurtma kodi<input name="order_code" placeholder="Masalan: AD-001"></label>
<label>Mijoz<input name="customer" placeholder="Mijoz ismi"></label>
<label class="full">Xarajat kategoriyasi<input name="category" placeholder="Masalan: MDF, Transport, Seh uchun"></label>
</div>
<div class="actions">
<button class="btn primary" type="submit">⬇ Excel hisobotni yuklab olish</button>
<a class="btn secondary" href="/">Bosh sahifa</a>
</div>
</form>
<div class="note">Avans va keyingi to‘lovlar ikki marta qo‘shilmaydi. Bu boshqaruv hisoboti; rasmiy soliq hisobotini buxgalter tekshiradi.</div>
</div></div>
<script>
function fmt(d){return d.toISOString().slice(0,10)}
function preset(kind){
  const now=new Date(), s=new Date(now), e=new Date(now);
  if(kind==='week'){const day=(now.getDay()+6)%7;s.setDate(now.getDate()-day)}
  if(kind==='month'){s.setDate(1)}
  if(kind==='last_month'){s.setMonth(now.getMonth()-1,1);e.setDate(0)}
  if(kind==='year'){s.setMonth(0,1)}
  document.getElementById('start').value=fmt(s);
  document.getElementById('end').value=fmt(e);
}
</script>
</body></html>
"""


def register_excel_module(app: Any, get_db: Callable[[], sqlite3.Connection]) -> None:
    """Moliya sahifasi va XLSX eksport yo‘llarini Flask ilovasiga ulaydi."""
    if MODULE_NAME in getattr(app, "blueprints", {}):
        return

    bp = Blueprint(MODULE_NAME, __name__)

    @bp.get("/moliya-hisobot")
    @_require_finance_access
    def finance_report_page() -> str:
        today = date.today()
        start, end = _month_bounds(today)
        return render_template_string(
            PAGE_HTML,
            start=start,
            end=end,
            export_url=url_for(f"{MODULE_NAME}.finance_report_xlsx"),
        )

    @bp.get("/moliya-hisobot.xlsx")
    @_require_finance_access
    def finance_report_xlsx() -> Response:
        today = date.today()
        default_start, default_end = _month_bounds(today)
        start = _safe_iso(request.args.get("start"), default_start)
        end = _safe_iso(request.args.get("end"), default_end)
        if start > end:
            start, end = end, start

        report = _collect_report(
            get_db=get_db,
            start=start,
            end=end,
            order_code=(request.args.get("order_code") or "").strip()[:80],
            customer=(request.args.get("customer") or "").strip()[:120],
            category=(request.args.get("category") or "").strip()[:120],
        )
        out = _build_workbook(report)

        # Audit yozuvi mavjud bo‘lsa, eksportni tarixga yozadi.
        try:
            conn = get_db()
            if _table_exists(conn, "audit_log"):
                cols = _columns(conn, "audit_log")
                if {"amal", "tafsilot"}.issubset(cols):
                    conn.execute(
                        "INSERT INTO audit_log(amal,tafsilot) VALUES(?,?)",
                        ("Excel moliyaviy hisobot", f"{start} — {end}"),
                    )
                    conn.commit()
            conn.close()
        except Exception:
            pass

        filename = f"Mebel360_moliyaviy_hisobot_{start}_{end}.xlsx"
        return send_file(
            out,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename,
        )

    app.register_blueprint(bp)

    @app.after_request
    def _moliya_excel_button(response: Response) -> Response:
        """Rahbar kirganida barcha HTML sahifalarga ixcham Excel tugmasini qo‘shadi."""
        try:
            if not _authorized():
                return response
            ctype = response.headers.get("Content-Type", "")
            if "text/html" not in ctype or response.direct_passthrough:
                return response
            html = response.get_data(as_text=True)
            if "</body>" not in html or "m360-excel-float" in html:
                return response
            button = """
            <a id="m360-excel-float" href="/moliya-hisobot"
               style="position:fixed;left:14px;bottom:14px;z-index:99998;
               background:#166534;color:white;text-decoration:none;padding:11px 14px;
               border-radius:12px;font:700 13px Arial;box-shadow:0 8px 25px rgba(15,23,42,.25)">
               📊 Excel hisobot
            </a>
            """
            response.set_data(html.replace("</body>", button + "</body>"))
            response.headers["Content-Length"] = str(len(response.get_data()))
        except Exception:
            return response
        return response
